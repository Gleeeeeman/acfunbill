import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta, timezone
import pytz
import msvcrt
import time

headers = {
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "zh-CH,zh;q=0.9",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
    "Connection": "keep-alive",
    }

def pwd_input():    
    chars = []   
    while True:  
        try:  
            newChar = msvcrt.getch().decode(encoding="utf-8")  
        except:  
            return input("你很可能不是在cmd命令行下运行，密码输入将不能隐藏:")  
        if newChar in '\r\n': # 如果是换行，则输入结束               
             break   
        elif newChar == '\b': # 如果是退格，则删除密码末尾一位并且删除一个星号   
             if chars:    
                 del chars[-1]   
                 msvcrt.putch('\b'.encode(encoding='utf-8')) # 光标回退一格  
                 msvcrt.putch( ' '.encode(encoding='utf-8')) # 输出一个空格覆盖原来的星号  
                 msvcrt.putch('\b'.encode(encoding='utf-8')) # 光标回退一格准备接受新的输入                   
        else:  
            chars.append(newChar)  
            msvcrt.putch('*'.encode(encoding='utf-8')) # 显示为星号  
    return (''.join(chars) )  

def getSession(username:str,password:str):
    data = {
        "username": username,
        "password": password,
        "key": "",
        "captcha": ""
        }
    url ="http://id.app.acfun.cn/rest/web/login/signin"
    response = session.post(url,headers = headers,data = data)
    responsedict = json.loads(response.text)
    return responsedict,response
    
def getGiftList(response):
    headers2 = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CH,zh;q=0.9",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
        "Connection": "keep-alive",
        "Content-Type" : "application/x-www-form-urlencoded",
        "origin" : "https://www.acfun.cn",
        "referer" : "https://www.acfun.cn/",
    }
    # get cookie
    cookie_value = ""
    for key,value in response.cookies.items():  
        cookie_value += key + '=' + value + ';'  
    headers2['cookie'] = cookie_value
    # get token
    tokenURL = "http://id.app.acfun.cn/rest/web/token/get"
    token = json.loads(session.post(tokenURL,headers = headers2 , data = {"sid" : "acfun.midground.api"}).text)['acfun.midground.api_st']
    # get gift list
    params = {
        "subBiz" : "mainApp",
        "kpn" : "ACFUN_APP",
        "kpf" : "PC_WEB",
        "acfun.midground.api_st" : token
    }
    giftListURL = "https://kuaishouzt.com/rest/zt/live/gift/all"
    giftListDict = json.loads(session.post(giftListURL,data = params).text)
    giftPriceList = {'桃子' : 1}
    for i in range(len(giftListDict['data']['giftList'])):
        if(giftListDict['data']['giftList'][i]['payWalletType'] == 1):
            giftPriceList[giftListDict['data']['giftList'][i]['giftName']] = giftListDict['data']['giftList'][i]['giftPrice']
            
    return giftPriceList

def getRewardRecords(pcursor:str):
    url = 'https://m.acfun.cn/rest/apph5-direct/pay/reward/giveRecords?pcursor='+ pcursor
    #获取一页打赏记录
    response = session.get(url,headers = headers)
    responsedict = json.loads(response.text)
    return responsedict
    
def totalRewardRecords():
    #创建excel表格
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "打赏统计"
    ws1.append(["**********************************************************************"])
    ws1.append(["截止至" + utc_currentTime.strftime("  %Y{}%m{}%d{} %H{}%M{}%S{}  ").format('年','月','日','时','分','秒') + '打赏礼物统计'])
    ws1.append(["**********************************************************************"])
    ws1.append([])
    ws1.append(['排名','UID','ID','AC币'])
    ws2 = wb.create_sheet("打赏明细")
    ws2.append(["**********************************************************************"])
    ws2.append(["截止至" + utc_currentTime.strftime("  %Y{}%m{}%d{} %H{}%M{}%S{}  ").format('年','月','日','时','分','秒') + '打赏礼物明细'])
    ws2.append(["**********************************************************************"])
    ws2.append([])
    ws2.append(['打赏时间（北京时间）','UID','ID','打赏礼物','打赏数量','AC币'])
    #调整表格格式
    for i in [1,2,3,4]:
        ws1.merge_cells('A' + str(i) + ':G' + str(i))
        ws1.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    for i in [1,2,3,4]:
        ws2.merge_cells('A' + str(i) + ':G' + str(i))
        ws2.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    ws1.column_dimensions['A'].width=13
    ws1.column_dimensions['B'].width=13
    ws1.column_dimensions['C'].width=28
    ws1.column_dimensions['D'].width=10
    ws2.column_dimensions['A'].width=31
    ws2.column_dimensions['B'].width=13
    ws2.column_dimensions['C'].width=28
    ws2.column_dimensions['D'].width=10
    ws2.column_dimensions['E'].width=10
    ws2.column_dimensions['F'].width=10
    for i in [2,4]:
        ws1.cell(row=5, column=i).alignment = Alignment(horizontal='right')
    for i in [1,2,5,6]:
        ws2.cell(row=5, column=i).alignment = Alignment(horizontal='right')
    #创建三个统计数组（打赏对象UID；打赏对象ID；总打赏AC币）
    UIDList = []
    IDList = []
    ACoinList = []
    pcursor = "0"   #获取数据参数初始化
    while pcursor != 'no_more':   #循环直到最后一页
        response = getRewardRecords(pcursor)
        if response['result'] != 0:     #获取信息失败重试
            continue
        else:                           #获取信息成功
            pcursor = response['pcursor']   #下一页的参数
            number = len(response['records'])
            for i in range(number):
                if (response['records'][i]['userId'] not in UIDList):   #添加打赏对象list
                    UIDList.append(response['records'][i]['userId'])
                    IDList.append(response['records'][i]['userName'])
                    ACoinList.append(0)
                #累计每个人的打赏AC币
                ACoinList[UIDList.index(response['records'][i]['userId'])] = ACoinList[UIDList.index(response['records'][i]['userId'])] + response['records'][i]['acoin']
                #由时间戳机算北京时间
                rewardTime = datetime.utcfromtimestamp(int(response['records'][i]['createTime'])/1000 + 28800).strftime("%Y{}%m{}%d{} %H{}%M{}%S{}  ").format('年','月','日','时','分','秒')
                #输出一行打赏明细到sheet2
                ws2.append([rewardTime,response['records'][i]['userId'],response['records'][i]['userName'],response['records'][i]['giftName'],response['records'][i]['giftCount'],response['records'][i]['acoin']])
    if len(UIDList) == 0:  #没有送礼记录的情况
        ws1.append(["无打赏记录"])
        ws2.append(["无打赏记录"])
    listLength = len(UIDList)
    totalTotal = 0
    while len(UIDList) != 0:    #打赏金额从高到低排序并输出到sheet1
        n = ACoinList.index(max(ACoinList))
        rank = listLength + 1 - len(UIDList)
        ws1.append([''.join(["第 ",str(rank)," 名"]),UIDList[n],IDList[n],ACoinList[n]])
        totalTotal = totalTotal + ACoinList[n]
        del UIDList[n]
        del IDList[n]
        del ACoinList[n]
    ws1.append(['','','合计：',totalTotal])
    #保存
    filename = "UID" + userID + "的打赏记录_" + currentTime + ".xlsx"
    print()
    try:
        wb.save(filename)
    except (PermissionError):
        print("文档保存失败，请关闭excel后重试")
        return True
    except:
        print("保存失败，请重试")
        return True
    print("数据获取完毕，已保存至" + filename)

def getDepositRecords(pcursor:str):
    url = 'https://m.acfun.cn/rest/apph5-direct/pay/bill/deposit/record?pcursor='+ pcursor
    #获取一页充值记录
    response = session.get(url,headers = headers)
    responsedict = json.loads(response.text)
    return responsedict

def totalDepositRecords():
    #创建excel表格
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "充值明细"
    ws1.append(["**********************************************************************"])
    ws1.append(["截止至" + utc_currentTime.strftime("  %Y{}%m{}%d{} %H{}%M{}%S{}  ").format('年','月','日','时','分','秒') + '充值记录'])
    ws1.append(["**********************************************************************"])
    ws1.append([])
    ws1.append(['充值时间','充值方式','充值金额','充值AC币数量'])
    #调整表格格式
    for i in [1,2,3,4]:
        ws1.merge_cells('A' + str(i) + ':D' + str(i))
        ws1.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    ws1.column_dimensions['A'].width=31
    ws1.column_dimensions['B'].width=15
    ws1.column_dimensions['C'].width=15
    ws1.column_dimensions['D'].width=15
    for i in [3,4]:
        ws1.cell(row=5, column=i).alignment = Alignment(horizontal='right')
    
    totalRMB = 0    #充值金额初始化
    totalACoin = 0  #充值AC币初始化
    pcursor = str(0)   #获取数据参数初始化
    while pcursor != 'no_more':   #循环直到最后一页
        response = getDepositRecords(pcursor)
        if response['result'] != 0:     #获取信息失败重试
            continue
        else:                           #获取信息成功
            pcursor = str(response['pcursor'])   #下一页的参数
            number = len(response['records'])
            for i in range(number):
                #累计充值AC币
                totalACoin = totalACoin + response['records'][i]['acoin']
                #累计充值金额
                totalRMB = float(float(totalACoin)/10)
                #由时间戳机算北京时间
                depositTime = datetime.utcfromtimestamp(int(response['records'][i]['createTime'])/1000 + 28800).strftime("%Y{}%m{}%d{} %H{}%M{}%S{}").format('年','月','日','时','分','秒')
                #输出一行打赏明细到sheet2
                ws1.append([depositTime,response['records'][i]['description'],float(float(response['records'][i]['acoin'])),response['records'][i]['acoin']])
    ws1.append(["总充值金额","",totalRMB,totalACoin])
    #保存
    filename = "UID" + userID + "的充值记录_" + currentTime + ".xlsx"
    print()
    try:
        wb.save(filename)
    except (PermissionError):
        print("文档保存失败，请关闭excel后重试")
        return True
    except:
        print("保存失败，请重试")
        return True
    print("数据获取完毕，已保存至" + filename)

def getReceiveRecords(pcursor:str):
    url = 'https://m.acfun.cn/rest/apph5-direct/pay/reward/receiveRecords?pcursor='+ pcursor
    #获取一页礼物记录
    response = session.get(url,headers = headers)
    responsedict = json.loads(response.text)
    return responsedict
    
def totalReceiveRecords():
    #创建excel表格
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "获得礼物统计"
    ws1.append(["******************************************************************************"])
    ws1.append(["截止至" + utc_currentTime.strftime("  %Y{}%m{}%d{} %H{}%M{}%S{}  ").format('年','月','日','时','分','秒') + '获得礼物统计'])
    ws1.append(["******************************************************************************"])
    ws1.append([])
    ws1.append(['排名','UID','ID','AC币','AC钻石'])
    ws2 = wb.create_sheet("获得礼物明细")
    ws2.append(["******************************************************************************"])
    ws2.append(["截止至" + utc_currentTime.strftime("  %Y{}%m{}%d{} %H{}%M{}%S{}  ").format('年','月','日','时','分','秒') + '获得礼物明细'])
    ws2.append(["******************************************************************************"])
    ws2.append([])
    ws2.append(['打赏时间（北京时间）','UID','ID','打赏礼物','打赏数量','AC币','AC钻石'])
    #调整表格格式
    for i in [1,2,3,4]:
        ws1.merge_cells('A' + str(i) + ':H' + str(i))
        ws1.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    for i in [1,2,3,4]:
        ws2.merge_cells('A' + str(i) + ':H' + str(i))
        ws2.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    ws1.column_dimensions['A'].width=13
    ws1.column_dimensions['B'].width=13
    ws1.column_dimensions['C'].width=28
    ws1.column_dimensions['D'].width=10
    ws1.column_dimensions['E'].width=10
    ws2.column_dimensions['A'].width=31
    ws2.column_dimensions['B'].width=13
    ws2.column_dimensions['C'].width=28
    ws2.column_dimensions['D'].width=10
    ws2.column_dimensions['E'].width=10
    ws2.column_dimensions['F'].width=10
    ws2.column_dimensions['G'].width=10
    for i in [2,4,5]:
        ws1.cell(row=5, column=i).alignment = Alignment(horizontal='right')
    for i in [2,5,6,7]:
        ws2.cell(row=5, column=i).alignment = Alignment(horizontal='right')
    #创建四个统计数组（打赏者UID；打赏者ID；总打赏AC币；总打赏AC钻石）
    UIDList = []
    IDList = []
    ACoinList = []
    azuanList = []
    pcursor = "0"   #获取数据参数初始化
    while pcursor != 'no_more':   #循环直到最后一页
        response = getReceiveRecords(pcursor)
        if response['result'] != 0:     #获取信息失败重试
            continue
        else:                           #获取信息成功
            pcursor = response['pcursor']   #下一页的参数
            number = len(response['records'])
            for i in range(number):
                if (response['records'][i]['userId'] not in UIDList):   #添加打赏者list
                    UIDList.append(response['records'][i]['userId'])
                    IDList.append(response['records'][i]['userName'])
                    ACoinList.append(0)
                    azuanList.append(0)
                #累计每个人的打赏AC钻石
                azuanList[UIDList.index(response['records'][i]['userId'])] = azuanList[UIDList.index(response['records'][i]['userId'])] + response['records'][i]['azuanAmount']
                #累计每个人的打赏AC币
                ACoinAmount = giftPriceList[response['records'][i]['giftName']] * response['records'][i]['giftCount']
                ACoinList[UIDList.index(response['records'][i]['userId'])] = ACoinList[UIDList.index(response['records'][i]['userId'])] + ACoinAmount
                #由时间戳机算北京时间
                receiveTime = datetime.utcfromtimestamp(int(response['records'][i]['createTime'])/1000 + 28800).strftime("%Y{}%m{}%d{} %H{}%M{}%S{}").format('年','月','日','时','分','秒')
                #输出一行打赏明细到sheet2
                ws2.append([receiveTime,response['records'][i]['userId'],response['records'][i]['userName'],response['records'][i]['giftName'],response['records'][i]['giftCount'],ACoinAmount,response['records'][i]['azuanAmount']])
    
    if len(UIDList) == 0:  #没有收礼记录的情况
        ws1.append(["无礼物记录"])
        ws2.append(["无礼物记录"])
    listLength = len(UIDList)
    totalTotalCoin = 0
    totalTotalZuan = 0
    while len(UIDList) != 0:    #打赏金额从高到低排序并输出到sheet1
        n = ACoinList.index(max(ACoinList))
        rank = listLength + 1 - len(UIDList)
        ws1.append([''.join(["第 ",str(rank)," 名"]),UIDList[n],IDList[n],ACoinList[n],azuanList[n]])
        totalTotalCoin = totalTotalCoin + ACoinList[n]
        totalTotalZuan = totalTotalZuan + azuanList[n]
        del UIDList[n]
        del IDList[n]
        del ACoinList[n]
        del azuanList[n]
    ws1.append(['','','合计：',totalTotalCoin,totalTotalZuan])
    #保存
    filename = "UID" + userID + "用户的获得礼物记录_" + currentTime + ".xlsx"
    print()
    try:
        wb.save(filename)
    except (PermissionError):
        print("文档保存失败，请关闭excel后重试")
        return True
    except:
        print("保存失败，请重试")
        return True
    print("数据获取完毕，已保存至" + filename)
    
def timeJudge(TIME):
    if len(TIME) not in [8,10,12,14]:
        return 0
    TIME = TIME.ljust(14,'1')
    if not (1 <= int(TIME[4:6]) <= 12):
        return 0
    if int(TIME[4:6]) in [1,3,5,7,8,10,12]:
        if not (1 <= int(TIME[6:8]) <= 31):
            return 0
    elif int(TIME[4:6]) in [4,6,9,11]:
        if not (1 <= int(TIME[6:8]) <= 30):
            return 0
    elif int(TIME[4:6]) == 2:
        if (int(TIME[0:4])%4 == 0 and int(TIME[0:4])%100 != 0) or int(TIME[0:4])%400 == 0:
            if not (1 <= int(TIME[6:8]) <= 29):
                return 0
        elif not (1 <= int(TIME[6:8]) <= 28):
            return 0
    if not (0 <= int(TIME[8:10]) <= 23):
        return 0
    if not (0 <= int(TIME[10:12]) < 59):
        return 0
    if not (0 <= int(TIME[12:14]) < 59):
        return 0
    return 1
    
def intervalReceiveRecords():
    #输入起始结束日期
    while True:
        print("请以【年年年年月月日日时时分分秒秒】的格式输入日期。（时分秒可省略）")
        print("例①：【20200807】即为2020年8月7日")
        print("例②：【2020080712】即为2020年8月7日12时")
        print("例③：【202008071230】即为2020年8月7日12时30分")
        print("如只输入至日期，则记录包括结束日期当天的记录")
        print("如需查询单日记录，则起始日期和结束日期为同一天")
        beginTime = input("请输入查询起始时间点：")
        endTime = input("请输入查询结束时间点：")
        if timeJudge(beginTime) == 0 or timeJudge(endTime) == 0:
            print()
            print("时间格式输入错误，请重新输入")
            print()
        elif beginTime.ljust(14,'0') > endTime.ljust(14,'0'):
            print()
            print("起始时间应早于结束时间，请重新输入")
            print()
        else:
            if len(endTime) == 8:
                endTime = str(int(endTime) * 1000000 + 235959)
            beginTimeArray = time.strptime(beginTime.ljust(14,'0'),"%Y%m%d%H%M%S")
            beginTimeStamp = time.mktime(pytz.timezone('Asia/Shanghai').localize(datetime.strptime(beginTime.ljust(14,'0'),"%Y%m%d%H%M%S")).astimezone(pytz.utc).utctimetuple()) - time.timezone
            beginTimeStamp = int(beginTimeStamp * 1000)
            endTimeArray = time.strptime(endTime.ljust(14,'0'),"%Y%m%d%H%M%S")
            endTimeStamp = time.mktime(pytz.timezone('Asia/Shanghai').localize(datetime.strptime(endTime.ljust(14,'0'),"%Y%m%d%H%M%S")).astimezone(pytz.utc).utctimetuple()) - time.timezone
            endTimeStamp = int(endTimeStamp * 1000)
            print()
            print("将查询" + time.strftime(" %Y{}%m{}%d{}%H{}%M{}%S{} ",beginTimeArray).format('年','月','日','时','分','秒') + "至" + time.strftime(" %Y{}%m{}%d{}%H{}%M{}%S{} ",endTimeArray).format('年','月','日','时','分','秒') + "的记录")
            print()
            print("确认吗？确认请按回车")
            confirm = input("重新输入请输入任意字符并按回车：")
            if confirm == "":
                break
            print()

    print("正在获取数据中，请稍后...")
    #创建excel表格
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "获得礼物统计"
    ws1.append(["******************************************************************************"])
    ws1.append([time.strftime("%Y{}%m{}%d{} %H{}%M{}%S{}  ",beginTimeArray).format('年','月','日','时','分','秒') + '至' + time.strftime("  %Y{}%m{}%d{} %H{}%M{}%S{}  ",endTimeArray).format('年','月','日','时','分','秒') + '获得礼物统计'])
    ws1.append(["******************************************************************************"])
    ws1.append([])
    ws1.append(['排名','UID','ID','AC币','AC钻石'])
    ws2 = wb.create_sheet("获得礼物明细")
    ws2.append(["******************************************************************************"])
    ws2.append([time.strftime("%Y{}%m{}%d{} %H{}%M{}%S{}  ",beginTimeArray).format('年','月','日','时','分','秒') + '至' + time.strftime("  %Y{}%m{}%d{} %H{}%M{}%S{}  ",endTimeArray).format('年','月','日','时','分','秒') + '获得礼物明细'])
    ws2.append(["******************************************************************************"])
    ws2.append([])
    ws2.append(['打赏时间（北京时间）','UID','ID','打赏礼物','打赏数量','AC币','AC钻石'])
    #调整表格格式
    for i in [1,2,3,4]:
        ws1.merge_cells('A' + str(i) + ':H' + str(i))
        ws1.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    for i in [1,2,3,4]:
        ws2.merge_cells('A' + str(i) + ':H' + str(i))
        ws2.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    ws1.column_dimensions['A'].width=13
    ws1.column_dimensions['B'].width=13
    ws1.column_dimensions['C'].width=28
    ws1.column_dimensions['D'].width=10
    ws1.column_dimensions['E'].width=10
    ws2.column_dimensions['A'].width=31
    ws2.column_dimensions['B'].width=13
    ws2.column_dimensions['C'].width=28
    ws2.column_dimensions['D'].width=10
    ws2.column_dimensions['E'].width=10
    ws2.column_dimensions['F'].width=10
    ws2.column_dimensions['G'].width=10
    for i in [2,4,5]:
        ws1.cell(row=5, column=i).alignment = Alignment(horizontal='right')
    for i in [2,5,6,7]:
        ws2.cell(row=5, column=i).alignment = Alignment(horizontal='right')
    #创建四个统计数组（打赏者UID；打赏者ID；总打赏AC币；总打赏AC钻石）
    UIDList = []
    IDList = []
    ACoinList = []
    azuanList = []
    pcursor = "0"   #获取数据参数初始化
    while pcursor != 'no_more':   #循环直到最后一页
        response = getReceiveRecords(pcursor)
        if response['result'] != 0:     #获取信息失败重试
            continue
        else:                           #获取信息成功
            pcursor = response['pcursor']   #下一页的参数
            number = len(response['records'])
            for i in range(number):
                if response['records'][i]['createTime'] > endTimeStamp:
                    continue
                if response['records'][i]['createTime'] < beginTimeStamp:
                    pcursor = 'no_more'
                    break
                if (response['records'][i]['userId'] not in UIDList):   #添加打赏者list
                    UIDList.append(response['records'][i]['userId'])
                    IDList.append(response['records'][i]['userName'])
                    ACoinList.append(0)
                    azuanList.append(0)
                #累计每个人的打赏AC钻石
                azuanList[UIDList.index(response['records'][i]['userId'])] = azuanList[UIDList.index(response['records'][i]['userId'])] + response['records'][i]['azuanAmount']
                #累计每个人的打赏AC币
                ACoinAmount = giftPriceList[response['records'][i]['giftName']] * response['records'][i]['giftCount']
                ACoinList[UIDList.index(response['records'][i]['userId'])] = ACoinList[UIDList.index(response['records'][i]['userId'])] + ACoinAmount
                #由时间戳机算北京时间
                receiveTime = datetime.utcfromtimestamp(int(response['records'][i]['createTime'])/1000 + 28800).strftime("%Y{}%m{}%d{} %H{}%M{}%S{}").format('年','月','日','时','分','秒')
                #输出一行打赏明细到sheet2
                ws2.append([receiveTime,response['records'][i]['userId'],response['records'][i]['userName'],response['records'][i]['giftName'],response['records'][i]['giftCount'],ACoinAmount,response['records'][i]['azuanAmount']])
    
    if len(UIDList) == 0:  #没有收礼记录的情况
        ws1.append(["此时间区间内无礼物记录"])
        ws2.append(["此时间区间内无礼物记录"])
    listLength = len(UIDList)
    totalTotalCoin = 0
    totalTotalZuan = 0
    while len(UIDList) != 0:    #打赏金额从高到低排序并输出到sheet1
        n = ACoinList.index(max(ACoinList))
        rank = listLength + 1 - len(UIDList)
        ws1.append([''.join(["第 ",str(rank)," 名"]),UIDList[n],IDList[n],ACoinList[n],azuanList[n]])
        totalTotalCoin = totalTotalCoin + ACoinList[n]
        totalTotalZuan = totalTotalZuan + azuanList[n]
        del UIDList[n]
        del IDList[n]
        del ACoinList[n]
        del azuanList[n]
    ws1.append(['','','合计：',totalTotalCoin,totalTotalZuan])
    #保存
    filename = "UID" + userID + "用户" + time.strftime("%Y%m%d%H%M%S",beginTimeArray) + '至' + time.strftime("%Y%m%d%H%M%S",endTimeArray) + '期间获得礼物记录' + ".xlsx"
    print()
    try:
        wb.save(filename)
    except (PermissionError):
        print("文档保存失败，请关闭excel后重试")
        return True
    except:
        print("保存失败，请重试")
        return True
    print("数据获取完毕，已保存至" + filename)

def lastMonthPeachRecords():
    if currentMonth == "01":
        lastMonth_year = str(int(currentYear) - 1)
        lastMonth_month = "12"
    else:
        lastMonth_year = currentYear
        lastMonth_month = str(int(currentMonth) - 1).rjust(2,'0')
    beginTime = ''.join([ lastMonth_year , lastMonth_month , "01"]).ljust(14,'0')
    endTime = ''.join([ currentYear , currentMonth , "01"]).ljust(14,'0')
    beginTimeStamp = time.mktime(pytz.timezone('Asia/Shanghai').localize(datetime.strptime(beginTime.ljust(14,'0'),"%Y%m%d%H%M%S")).astimezone(pytz.utc).utctimetuple()) - time.timezone
    beginTimeStamp = int(beginTimeStamp * 1000)
    endTimeStamp = time.mktime(pytz.timezone('Asia/Shanghai').localize(datetime.strptime(endTime.ljust(14,'0'),"%Y%m%d%H%M%S")).astimezone(pytz.utc).utctimetuple()) - time.timezone
    endTimeStamp = int(endTimeStamp * 1000)
    print("正在获取上月桃榜数据中，请稍后...")
    #创建excel表格
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "上月获得桃子统计"
    ws1.append(["******************************************************************"])
    ws1.append([lastMonth_year + "年" + lastMonth_month + "月 " + '获得桃子统计'])
    ws1.append(["******************************************************************"])
    ws1.append([])
    ws1.append(['排名','UID','ID','打赏桃子数量'])
    ws2 = wb.create_sheet("上月获得桃子明细")
    ws2.append(["******************************************************************"])
    ws2.append([lastMonth_year + "年" + lastMonth_month + "月 " + '获得桃子明细'])
    ws2.append(["******************************************************************"])
    ws2.append([])
    ws2.append(['打赏时间（北京时间）','UID','ID','打赏礼物','打赏数量'])
    #调整表格格式
    for i in [1,2,3,4]:
        ws1.merge_cells('A' + str(i) + ':E' + str(i))
        ws1.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    for i in [1,2,3,4]:
        ws2.merge_cells('A' + str(i) + ':E' + str(i))
        ws2.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    ws1.column_dimensions['A'].width=13
    ws1.column_dimensions['B'].width=13
    ws1.column_dimensions['C'].width=28
    ws1.column_dimensions['D'].width=20
    ws2.column_dimensions['A'].width=31
    ws2.column_dimensions['B'].width=13
    ws2.column_dimensions['C'].width=28
    for i in [2,4]:
        ws1.cell(row=5, column=i).alignment = Alignment(horizontal='right')
    for i in [2,5]:
        ws2.cell(row=5, column=i).alignment = Alignment(horizontal='right')
    #创建三个统计数组（打赏者UID；打赏者ID；总打赏桃子数量）
    UIDList = []
    IDList = []
    PeachList = []
    pcursor = "0"   #获取数据参数初始化
    while pcursor != 'no_more':   #循环直到最后一页
        response = getReceiveRecords(pcursor)
        if response['result'] != 0:     #获取信息失败重试
            continue
        else:                           #获取信息成功
            pcursor = response['pcursor']   #下一页的参数
            number = len(response['records'])
            for i in range(number):
                if response['records'][i]['createTime'] > endTimeStamp:
                    continue
                if response['records'][i]['createTime'] < beginTimeStamp:
                    pcursor = 'no_more'
                    break
                if response['records'][i]['giftName'] != "桃子":
                    continue
                if (response['records'][i]['userId'] not in UIDList):   #添加打赏者list
                    UIDList.append(response['records'][i]['userId'])
                    IDList.append(response['records'][i]['userName'])
                    PeachList.append(0)
                #累计每个人的打赏桃子数量
                PeachList[UIDList.index(response['records'][i]['userId'])] = PeachList[UIDList.index(response['records'][i]['userId'])] + response['records'][i]['giftCount']
                #由时间戳机算北京时间
                receiveTime = datetime.utcfromtimestamp(int(response['records'][i]['createTime'])/1000 + 28800).strftime("%Y{}%m{}%d{} %H{}%M{}%S{}").format('年','月','日','时','分','秒')
                #输出一行打赏明细到sheet2
                ws2.append([receiveTime,response['records'][i]['userId'],response['records'][i]['userName'],response['records'][i]['giftName'],response['records'][i]['giftCount']])
    
    if len(UIDList) == 0:  #没有收礼记录的情况
        ws1.append(["上个月无获得桃子记录"])
        ws2.append(["上个月无获得桃子记录"])
    listLength = len(UIDList)
    totalTotal = 0
    while len(UIDList) != 0:    #打赏桃子数量从高到低排序并输出到sheet1
        n = PeachList.index(max(PeachList))
        rank = listLength + 1 - len(UIDList)
        ws1.append([''.join(["第 ",str(rank)," 名"]),UIDList[n],IDList[n],PeachList[n]])
        totalTotal = totalTotal + PeachList[n]
        del UIDList[n]
        del IDList[n]
        del PeachList[n]
    ws1.append(['','','合计：',totalTotal])
    #保存
    filename = "UID" + userID + "用户" + lastMonth_year + "年" + lastMonth_month + "月获得桃子记录" + ".xlsx"
    print()
    try:
        wb.save(filename)
    except (PermissionError):
        print("文档保存失败，请关闭excel后重试")
        return True
    except:
        print("保存失败，请重试")
        return True
    print("数据获取完毕，已保存至" + filename)
    
    
def updateRemind():
    response = requests.get(url)
    responsedict = json.loads(response.text)
    
    versionUrl = responsedict['versionUrl']
    downloadUrl = responsedict['downloadUrl'] 
    
    response = requests.get(versionUrl)
    responsedict = json.loads(response.text)
    versionNumber = len(responsedict['records'])
    if float(version) < float(responsedict['records'][versionNumber - 1]['version']):
        print("=====================================")
        print("此工具当前版本为V" + version + "，最新版本为V" + responsedict['records'][versionNumber - 1]['version'] + "，更新内容如下：")
        for i in range(versionNumber):
            if float(version) < float(responsedict['records'][i]['version']):
                print("=====================================")
                print("版本：V" + responsedict['records'][i]['version'])
                print("更新日期：" + responsedict['records'][i]['updateTime'])
                print("更新内容：" + responsedict['records'][i]['content'])
        print("=====================================")
        print("如需更新，请登录   " + downloadUrl + "  下载最新版本")
        print("=====================================")
        input("按回车键继续使用当前版本")
        print()
        print()
    return True
    
if __name__ == "__main__":
    version = "4.1"
    url = 'http://raw.githubusercontent.com/Gleeeeeman/acfunbill/main/url.json'
    versionUrl = 'http://raw.githubusercontent.com/Gleeeeeman/acfunbill/main/version.json'
    downloadUrl = 'https://github.com/Gleeeeeman/acfunbill/releases/'
    try:
        updateRemind()
    except:
        print("获取版本更新信息失败")
        print("将使用  "+ version +"  版本")
        print("手动检查更新请登录  " + downloadUrl)
    
    print()
    print('======AcFun打赏充值礼物记录查询小工具V' + version + '___By游学志========')
    print()
    print('==================更新日期：2021年09月12日==================')
    print()
    print('============本工具所有时间以北京时间(UTC+8)为准=============')
    print()
    print('==========================登录AcFun=========================')
    utc_currentTime = datetime.utcnow().replace(tzinfo=timezone.utc)
    currentTime = utc_currentTime.astimezone(timezone(timedelta(hours=8))).strftime('%Y%m%d')
    currentYear = utc_currentTime.astimezone(timezone(timedelta(hours=8))).strftime('%Y')
    currentMonth = utc_currentTime.astimezone(timezone(timedelta(hours=8))).strftime('%m')
    while True:
        print('请输入手机号/邮箱')
        username = str(input())
        print('请输入密码')
        password = str(pwd_input())
        print()
        session = requests.Session()
        responsedict,response = getSession(username,password)
        if responsedict['result'] != 0:
            print()
            print(responsedice['error_msg'])
            print('请检查登录信息并重试')
            print()
            continue
        else:
            # get gift price list
            giftPriceList = getGiftList(response)
            print()
            print("|==========================================================|")
            print("||                                                        ||")
            print("||                        登录成功                        ||")
            print("||                                                        ||")
            print("|==========================================================|")
            userID = str(responsedict['userId'])
            username = responsedict['username']
            print()
            print("     用户名：" + username + "     UID：" + userID)
            break
            
    while True:
        print()
        print("============================================================")
        print("获取打赏记录请输入1")
        print("获取充值记录请输入2")
        print("获取收到礼物记录请输入3")
        print("获取指定时间区间礼物记录请输入4")
        print("获取上个月桃榜记录请输入5")
        print("退出请输入9")
        feature = input("请输入序号并按回车键：")
        print()
        if feature == "1":
            print("正在获取数据中，请稍后...")
            totalRewardRecords()
            continue
        elif feature == "2":
            print("正在获取数据中，请稍后...")
            totalDepositRecords()
            continue
        elif feature == "3":
            print("正在获取数据中，请稍后...")
            totalReceiveRecords()
            continue
        elif feature == "4":
            intervalReceiveRecords()
            continue
        elif feature == "5":
            lastMonthPeachRecords()
            continue
        elif feature == "9":
            break
        else:
            print()
            print("输入编号错误，请重新输入")
    
    print()
    print("==========================感谢使用==========================")
    input("回车键退出")